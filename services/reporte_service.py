# ============================================================
# services/reporte_service.py — Generación de Reportes PDF y Excel
# ============================================================

import io
from datetime import datetime
from config import EMPRESA

# ---- ReportLab (PDF) ----
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ---- OpenPyXL (Excel) ----
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter


# =============================================================
# HELPERS
# =============================================================

def _fmt_fecha(dt) -> str:
    """Formatea un datetime o retorna '—'."""
    if dt:
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    return "—"


def _fmt_peso(val) -> str:
    """Formatea un peso o retorna '—'."""
    if val is not None:
        return f"{float(val):,.2f} KG"
    return "—"


def _fmt_solo_fecha(dt) -> str:
    """Solo la fecha (dd/mm/aaaa), sin hora."""
    return dt.strftime("%d/%m/%Y") if dt else "—"


def _fmt_solo_hora(dt) -> str:
    """Solo la hora, en formato 12h con AM/PM (igual al ticket de referencia)."""
    return dt.strftime("%I:%M:%S %p") if dt else "—"


# Traduce tipo_pesaje al lenguaje de "Operación" del formato de ticket
# heredado: pesaje general = mercancía que entra a la planta
# (recepción), producto terminado = mercancía que sale (despacho).
_OPERACION_POR_TIPO = {
    "GENERAL": "RECEPCIÓN",
    "PRODUCTO_TERMINADO": "DESPACHO",
}


# =============================================================
# TICKET DE PESAJE (PDF)
# =============================================================

def generar_ticket_pdf(pesada) -> bytes:
    """
    Genera el ticket de pesaje en PDF con el formato de planilla que ya
    usaba la empresa con el sistema anterior (Bigsoft) -- mismo layout
    y campos, alimentado con los datos de Pesada en vez de texto fijo,
    para que operadores y choferes no tengan que aprender un documento
    nuevo. Los datos de la empresa salen de config.py (EMPRESA), no
    están hardcodeados acá.

    Args:
        pesada: Objeto Pesada de la base de datos

    Returns:
        Contenido del PDF en bytes (Romana y Centro de Costos son
        máquinas distintas sin disco compartido, así que el reporte se
        genera en el servidor y se transmite por HTTP en vez de
        escribirse a un archivo local).
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.2*cm, bottomMargin=1.2*cm
    )
    elements = []
    NEGRO = colors.black
    ANCHO_TOTAL = 18*cm

    style_empresa = ParagraphStyle("empresa", fontName="Courier-Bold", fontSize=11,
                                    alignment=TA_CENTER, leading=13)
    style_empresa_sub = ParagraphStyle("empresa_sub", fontName="Courier", fontSize=7.5,
                                        alignment=TA_CENTER, leading=10)
    style_lbl = ParagraphStyle("lbl", fontName="Courier", fontSize=9, leading=13)
    style_ticket_nro = ParagraphStyle("ticket_nro", fontName="Courier-Bold", fontSize=13,
                                       alignment=TA_RIGHT, leading=15)
    style_fecha_hdr = ParagraphStyle("fecha_hdr", fontName="Courier", fontSize=9,
                                      alignment=TA_RIGHT, leading=12)

    # ---- Encabezado: empresa a la izq., ticket nro/fecha a la der. ----
    celda_empresa = [
        Paragraph(EMPRESA["nombre"], style_empresa),
        Paragraph(EMPRESA["direccion"], style_empresa_sub),
        Paragraph(f"VENEZUELA. TELF.{EMPRESA['telefono']}", style_empresa_sub),
    ]
    celda_ticket = [
        Paragraph(f"TICKET NRO: <b>{pesada.numero_ticket}</b>", style_ticket_nro),
        Paragraph(f"Fecha: {_fmt_solo_fecha(pesada.fecha_salida or pesada.fecha_entrada)}",
                  style_fecha_hdr),
    ]
    tabla_header = Table([[celda_empresa, celda_ticket]], colWidths=[12*cm, 6*cm])
    tabla_header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(tabla_header)
    elements.append(Spacer(1, 4))
    elements.append(HRFlowable(width="100%", thickness=1, color=NEGRO))
    elements.append(Spacer(1, 2))

    elements.append(Paragraph(
        f"Telefono: &nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Fax: &nbsp;&nbsp;&nbsp;&nbsp;&nbsp; "
        f"RIF: {EMPRESA['rif']}", style_lbl))
    elements.append(Spacer(1, 4))
    elements.append(HRFlowable(width="100%", thickness=1, color=NEGRO))
    elements.append(Spacer(1, 6))

    # ---- Datos de la operación: cliente, producto, chofer, vehículo... ----
    proveedor_txt = (f"{pesada.proveedor.codigo} {pesada.proveedor.nombre}"
                      if pesada.proveedor else (pesada.empresa_cliente_proveedor or "—"))
    producto_txt = (f"{pesada.producto.codigo} {pesada.producto.nombre}"
                     if pesada.producto else "—")
    if pesada.conductor:
        chofer_txt = f"{pesada.conductor.documento} &nbsp;&nbsp; {pesada.conductor.nombre}"
    else:
        chofer_txt = pesada.cedula_conductor_libre or "—"
    placa_txt = pesada.vehiculo.placa if pesada.vehiculo else "—"
    remolque_txt = pesada.remolque.placa if pesada.remolque else ""
    operacion_txt = _OPERACION_POR_TIPO.get(pesada.tipo_pesaje, pesada.tipo_pesaje)
    transporte_txt = (pesada.transportista.nombre if pesada.transportista
                       else (pesada.empresa_transportista or "—"))
    orden_txt = pesada.orden_compra or "—"
    cantidad_txt = f"{float(pesada.cantidad):,.2f}" if pesada.cantidad else "0.00"

    filas_info = [
        [Paragraph(f"<b>CLIENTE</b>&nbsp;&nbsp;&nbsp;: {proveedor_txt}", style_lbl),
         Paragraph(f"Producto: {producto_txt}", style_lbl)],
        [Paragraph(f"Chofer C.I. {chofer_txt}", style_lbl),
         Paragraph(f"Placa: {placa_txt} &nbsp;&nbsp;&nbsp; Remolque: {remolque_txt}", style_lbl)],
        [Paragraph("Procedencia:", style_lbl),
         Paragraph(f"Operacion: {operacion_txt}", style_lbl)],
        [Paragraph(f"Transporte: {transporte_txt}", style_lbl),
         Paragraph(f"Orden de Compra: {orden_txt} &nbsp;&nbsp;&nbsp; Cantidad: {cantidad_txt}",
                    style_lbl)],
    ]
    tabla_info = Table(filas_info, colWidths=[ANCHO_TOTAL/2, ANCHO_TOTAL/2])
    tabla_info.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(tabla_info)
    elements.append(Spacer(1, 8))

    # ---- Tabla de pesadas: Tara / Peso Bruto / Peso Final ----
    # Peso Final es el 3er pesaje (nuevo en este sistema, antes de
    # autorizar la salida) -- se agrega como 3ra columna solo si existe,
    # las pesadas completadas antes de introducir este campo no lo tienen.
    columnas_peso = [
        ("TARA", pesada.fecha_entrada, pesada.peso_tara),
        ("PESO BRUTO", pesada.fecha_captura, pesada.peso_bruto),
    ]
    if pesada.peso_final is not None:
        columnas_peso.append(("PESO FINAL", pesada.fecha_salida, pesada.peso_final))

    style_peso_hdr = ParagraphStyle("peso_hdr", fontName="Courier-Bold", fontSize=9,
                                     alignment=TA_CENTER)
    style_peso_val = ParagraphStyle("peso_val", fontName="Courier-Bold", fontSize=11,
                                     alignment=TA_CENTER)
    style_peso_lbl = ParagraphStyle("peso_lbl", fontName="Courier", fontSize=8,
                                     alignment=TA_CENTER)

    fila_encabezados = [Paragraph(nombre, style_peso_hdr) for nombre, _, _ in columnas_peso]
    fila_fecha = [Paragraph(f"Fecha: {_fmt_solo_fecha(f)}", style_peso_lbl)
                  for _, f, _ in columnas_peso]
    fila_hora = [Paragraph(f"Hora: {_fmt_solo_hora(f)}", style_peso_lbl)
                 for _, f, _ in columnas_peso]
    fila_peso = [Paragraph(f"{float(p or 0):,.2f} KGS", style_peso_val)
                 for _, _, p in columnas_peso]

    ancho_col = ANCHO_TOTAL / len(columnas_peso)
    tabla_pesadas = Table(
        [fila_encabezados, fila_fecha, fila_hora, fila_peso],
        colWidths=[ancho_col] * len(columnas_peso)
    )
    tabla_pesadas.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, NEGRO),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tabla_pesadas)
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(f"Precintos de Seguridad: {pesada.precintos or '—'}", style_lbl))
    if pesada.observaciones:
        elements.append(Paragraph(f"Observaciones: {pesada.observaciones}", style_lbl))
    elements.append(Spacer(1, 8))

    # ---- Peso Neto + diferencia contra la cantidad declarada ----
    diferencia_txt = "—"
    if pesada.cantidad and float(pesada.cantidad) != 0:
        dif_pct = ((float(pesada.peso_neto or 0) - float(pesada.cantidad))
                   / float(pesada.cantidad) * 100)
        diferencia_txt = f"{dif_pct:,.2f} %"

    style_dif_lbl = ParagraphStyle("dif_lbl", fontName="Courier", fontSize=9)
    style_neto_lbl = ParagraphStyle("neto_lbl", fontName="Courier-Bold", fontSize=10)
    style_neto_val = ParagraphStyle("neto_val", fontName="Courier-Bold", fontSize=16,
                                     alignment=TA_CENTER)

    celda_dif = [
        Paragraph("Diferencia c/ Cantidad:", style_dif_lbl),
        Paragraph(diferencia_txt, style_dif_lbl),
    ]
    celda_neto = [
        Paragraph("PESO NETO", style_neto_lbl),
        Paragraph(f"{float(pesada.peso_neto or 0):,.2f} KGS", style_neto_val),
    ]
    tabla_neto = Table([[celda_dif, celda_neto]], colWidths=[ANCHO_TOTAL/2, ANCHO_TOTAL/2])
    tabla_neto.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, NEGRO),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(tabla_neto)
    elements.append(Spacer(1, 30))

    # ---- Firmas ----
    usr_firma = (pesada.usuario_salida.nombre_completo if pesada.usuario_salida
                 else (pesada.usuario_entrada.nombre_completo if pesada.usuario_entrada else "—"))
    conductor_firma = (pesada.conductor.nombre if pesada.conductor
                        else (pesada.cedula_conductor_libre or "—"))

    style_firma_lbl = ParagraphStyle("firma_lbl", fontName="Courier", fontSize=9)
    style_firma_nom = ParagraphStyle("firma_nom", fontName="Courier-Bold", fontSize=9,
                                      alignment=TA_CENTER)

    tabla_firmas = Table([
        [Paragraph("Romanero:", style_firma_lbl), "",
         Paragraph("Conductor:", style_firma_lbl), ""],
        ["", Paragraph(usr_firma, style_firma_nom),
         "", Paragraph(conductor_firma, style_firma_nom)],
    ], colWidths=[2.3*cm, 6.7*cm, 2.5*cm, 6.5*cm])
    tabla_firmas.setStyle(TableStyle([
        ("LINEABOVE", (1, 0), (1, 0), 1, NEGRO),
        ("LINEABOVE", (3, 0), (3, 0), 1, NEGRO),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(tabla_firmas)

    # ---- Construir PDF ----
    doc.build(elements)
    return buffer.getvalue()


# =============================================================
# KARDEX DE PESADAS (PDF)
# =============================================================

def generar_kardex_pdf(pesadas: list, titulo: str = "Kardex de Pesadas") -> bytes:
    """
    Genera un reporte PDF del Kardex con la lista de pesadas.

    Args:
        pesadas: Lista de objetos Pesada
        titulo:  Título del reporte

    Returns:
        Contenido del PDF en bytes.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    elements = []

    AZUL = colors.HexColor("#1E3A5F")
    VERDE = colors.HexColor("#00703C")

    # Encabezado
    style_h = ParagraphStyle("h", parent=styles["Normal"],
        fontSize=14, fontName="Helvetica-Bold",
        alignment=TA_CENTER, textColor=AZUL)
    style_sub = ParagraphStyle("sub", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica",
        alignment=TA_CENTER, textColor=colors.grey)

    elements.append(Paragraph(EMPRESA["nombre"], style_h))
    elements.append(Paragraph(titulo, style_h))
    elements.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(pesadas)} registros",
        style_sub))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=2, color=AZUL))
    elements.append(Spacer(1, 8))

    # Tabla de datos
    encabezados = ["Ticket", "Fecha", "Placa", "Producto", "Bruto KG", "Tara KG", "Neto KG", "Estado"]

    filas = [encabezados]
    total_neto = 0.0

    for p in pesadas:
        neto = float(p.peso_neto or 0)
        total_neto += neto
        filas.append([
            p.numero_ticket,
            _fmt_fecha(p.fecha_entrada)[:10],
            p.vehiculo.placa if p.vehiculo else "—",
            (p.producto.nombre[:12] if p.producto else "—"),
            f"{float(p.peso_bruto or 0):,.0f}",
            f"{float(p.peso_tara or 0):,.0f}",
            f"{neto:,.0f}",
            p.estado.upper()
        ])

    # Fila de totales
    filas.append([
        "", "TOTALES", "", "", "", "", f"{total_neto:,.0f}", f"{len(pesadas)} pesadas"
    ])

    col_widths = [2.8*cm, 2.5*cm, 2*cm, 3*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.1*cm]

    tabla = Table(filas, colWidths=col_widths, repeatRows=1)
    tabla.setStyle(TableStyle([
        # Encabezado
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 8),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),

        # Cuerpo
        ("FONTNAME",   (0, 1), (-1, -2), "Helvetica"),
        ("FONTSIZE",   (0, 1), (-1, -2), 8),
        ("ALIGN",      (4, 1), (-1, -1), "RIGHT"),

        # Fila de totales
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
        ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, -1), (-1, -1), 9),

        # Alternado
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9F9F9")]),

        # Bordes
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("BOX",        (0, 0), (-1, -1), 1.5, AZUL),

        # Padding
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
    ]))

    elements.append(tabla)
    doc.build(elements)
    return buffer.getvalue()


# =============================================================
# KARDEX EN EXCEL
# =============================================================

def generar_kardex_excel(pesadas: list, titulo: str = "Kardex de Pesadas") -> bytes:
    """
    Genera el Kardex en formato Excel (.xlsx).

    Args:
        pesadas: Lista de objetos Pesada
        titulo:  Título del reporte

    Returns:
        Contenido del archivo Excel en bytes.
    """
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    # ---- Estilos ----
    AZUL_OSCURO = "1E3A5F"
    AZUL_CLARO  = "1E90FF"
    VERDE       = "00703C"
    GRIS_CLARO  = "F5F5F5"
    GRIS_MEDIO  = "CCCCCC"

    font_titulo   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitulo= Font(name="Calibri", size=11, color="666666")
    font_header   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_normal   = Font(name="Calibri", size=10)
    font_total    = Font(name="Calibri", size=11, bold=True)

    fill_header = PatternFill("solid", fgColor=AZUL_OSCURO)
    fill_total  = PatternFill("solid", fgColor="E8F5E9")
    fill_fila1  = PatternFill("solid", fgColor="FFFFFF")
    fill_fila2  = PatternFill("solid", fgColor=GRIS_CLARO)

    borde = Border(
        left=Side(style="thin", color=GRIS_MEDIO),
        right=Side(style="thin", color=GRIS_MEDIO),
        top=Side(style="thin", color=GRIS_MEDIO),
        bottom=Side(style="thin", color=GRIS_MEDIO)
    )

    # ---- Fila 1: Nombre empresa ----
    ws.merge_cells("A1:J1")
    ws["A1"] = EMPRESA["nombre"]
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=AZUL_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- Fila 2: Título del reporte ----
    ws.merge_cells("A2:J2")
    ws["A2"] = titulo
    ws["A2"].font = Font(name="Calibri", size=13, bold=True, color=AZUL_CLARO)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # ---- Fila 3: Fecha de generación ----
    ws.merge_cells("A3:J3")
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(pesadas)} registros"
    ws["A3"].font = font_subtitulo
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18

    # ---- Fila 5: Encabezados de columnas ----
    columnas = [
        ("A", "TICKET",       15),
        ("B", "FECHA ENTRADA",20),
        ("C", "FECHA SALIDA", 20),
        ("D", "PLACA",        12),
        ("E", "CONDUCTOR",    22),
        ("F", "PRODUCTO",     22),
        ("G", "PROVEEDOR",    22),
        ("H", "BRUTO (KG)",   14),
        ("I", "TARA (KG)",    14),
        ("J", "NETO (KG)",    14),
    ]

    for col, nombre, ancho in columnas:
        celda = ws[f"{col}5"]
        celda.value = nombre
        celda.font = font_header
        celda.fill = fill_header
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde
        ws.column_dimensions[col].width = ancho

    ws.row_dimensions[5].height = 28

    # ---- Datos ----
    total_neto = 0.0
    for i, p in enumerate(pesadas, start=6):
        neto = float(p.peso_neto or 0)
        total_neto += neto
        fill_fila = fill_fila1 if i % 2 == 0 else fill_fila2

        valores = [
            p.numero_ticket,
            _fmt_fecha(p.fecha_entrada),
            _fmt_fecha(p.fecha_salida),
            p.vehiculo.placa if p.vehiculo else "—",
            p.conductor.nombre if p.conductor else "—",
            p.producto.nombre if p.producto else "—",
            p.proveedor.nombre if p.proveedor else "—",
            float(p.peso_bruto or 0),
            float(p.peso_tara or 0),
            neto,
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=i, column=col_idx, value=valor)
            celda.font = font_normal
            celda.fill = fill_fila
            celda.border = borde
            # Alinear pesos a la derecha
            if col_idx >= 8:
                celda.alignment = Alignment(horizontal="right")
                celda.number_format = "#,##0.00"

    # ---- Fila de totales ----
    fila_total = len(pesadas) + 6
    ws.cell(row=fila_total, column=1, value="TOTAL").font = font_total
    ws.cell(row=fila_total, column=10, value=total_neto).font = Font(
        name="Calibri", size=12, bold=True, color=VERDE)
    ws.cell(row=fila_total, column=10).number_format = "#,##0.00"

    for col_idx in range(1, 11):
        celda = ws.cell(row=fila_total, column=col_idx)
        celda.fill = fill_total
        celda.border = borde

    # ---- Congelar encabezado ----
    ws.freeze_panes = "A6"

    wb.save(buffer)
    return buffer.getvalue()
